import openpyxl 
import smtplib 
from email.mime.multipart import MIMEMultipart 
from email.mime.text import MIMEText 

# loading the excel sheet 
book = openpyxl.load_workbook('data\\attendance.xlsx') 

# Choose the sheet 
sheet = book['Sheet1'] 

# counting number of rows / students 
r = sheet.max_row 

# variable for looping for input 
resp = 1

# counting number of columns / subjects 
c = sheet.max_column 

# list of students to remind 
l1 = [] 

# to concatenate list of roll numbers with 
# lack of attendance 
l2 = "" 

# list of roll numbers with lack of attendance 
l3 = [] 

# staff mail ids 
staff_mails = ['erakshaya485@gmail.com', 'yyyyyyyy@gmail.com'] 

# Warning messages 
m1 = "warning!!! you can take only one more day leave for CI class"
m2 = "warning!!! you can take only one more day leave for python class"
m3 = "warning!!! you can take only one more day leave for DM class"


def savefile(): 
	book.save(r'D:\\attendance.xlsx') 
	print("saved!") 


def check(no_of_days, row_num, b): 

	# to use the globally declared lists and strings 
	global staff_mails 
	global l2 
	global l3 

	for student in range(0, len(row_num)): 
		# if total no.of.leaves equals threshold 
		if no_of_days[student] == 2: 
			if b == 1: 
				
				# mail_id appending 
				l1.append(sheet.cell(row=row_num[student], column=2).value) 
				mailstu(l1, m1) # sending mail 
			elif b == 2: 
				l1.append(sheet.cell(row=row_num[student], column=2).value) 
				mailstu(l1, m2) 
			else: 
				l1.append(sheet.cell(row=row_num[student], column=2).value) 
				mailstu(l1, m3) 

		# if total.no.of.leaves > threshold 
		elif no_of_days[student] > 2: 
			if b == 1: 

				# adding roll no 
				l2 = l2+str(sheet.cell(row=row_num[student], column=1).value) 

				# student mail_id appending 
				l3.append(sheet.cell(row=row_num[student], column=2).value) 
				subject = "CI" # subject based on the code number 

			elif b == 2: 
				l2 = l2+str(sheet.cell(row=row_num[student], column=1).value) 
				l3.append(sheet.cell(row=row_num[student], column=2).value) 
				subject = "Python"

			else: 
				l2 = l2+str(sheet.cell(row=row_num[student], column=1).value) 
				l3.append(sheet.cell(row=row_num[student], column=2).value) 
				subject = "Data mining"

		# If threshold crossed, modify the message 
		if l2 != "" and len(l3) != 0: 

			# message for student 
			msg1 = "you have lack of attendance in " + subject + " !!!"

			# message for staff 
			msg2 = "the following students have lack of attendance in your subject : "+l2 

			mailstu(l3, msg1) # mail to students 
			staff_id = staff_mails[b-1] # pick respective staff's mail_id 
			mailstaff(staff_id, msg2) # mail to staff 

# for students 
def mailstu(li, msg): 
	from_id = 'crazygirlaks@gmail.com'
	pwd = 'ERAkshaya485'
	s = smtplib.SMTP('smtp.gmail.com', 587, timeout=120) 
	s.starttls() 
	s.login(from_id, pwd) 

	# for each student to warn send mail 
	for i in range(0, len(li)): 
		to_id = li[i] 
		message = MIMEMultipart() 
		message['Subject'] = 'Attendance report'
		message.attach(MIMEText(msg, 'plain')) 
		content = message.as_string() 
		s.sendmail(from_id, to_id, content) 
		s.quit() 
	print("mail sent to students") 

# for staff 
def mailstaff(mail_id, msg): 
	from_id = 'crazygirlaks@gmail.com'
	pwd = 'ERAkshaya485'
	to_id = mail_id 
	message = MIMEMultipart() 
	message['Subject'] = 'Lack of attendance report'
	message.attach(MIMEText(msg, 'plain')) 
	s = smtplib.SMTP('smtp.gmail.com', 587, timeout=120) 
	s.starttls() 
	s.login(from_id, pwd) 
	content = message.as_string() 
	s.sendmail(from_id, to_id, content) 
	s.quit() 
	print('Mail Sent to staff') 


while resp == 1: 
	print("1--->CI\n2--->python\n3--->DM") 

	# enter the correspondingnumber 
	y = int(input("enter subject :")) 

	# no.of.absentees for that subject 
	no_of_absentees = int(input('no.of.absentees :')) 

	if(no_of_absentees > 1): 
		x = list(map(int, (input('roll nos :').split(' ')))) 
	else: 
		x = [int(input('roll no :'))] 

	# list to hold row of the student in Excel sheet 
	row_num = [] 

	# list to hold total no.of leaves 
	# taken by ith student 
	no_of_days = [] 

	for student in x: 

		for i in range(2, r+1): 

			if y == 1: 
				if sheet.cell(row=i, column=1).value == student: 
					m = sheet.cell(row=i, column=3).value 
					m = m+1
					sheet.cell(row=i, column=3).value = m 
					savefile() 
					no_of_days.append(m) 
					row_num.append(i) 

			elif y == 2: 
				if sheet.cell(row=i, column=1).value == student: 
					m = sheet.cell(row=i, column=4).value 
					m = m+1
					sheet.cell(row=i, column=4).value = m+1
					no_of_days.append(m) 
					row_num.append(i) 

			elif y ==  3: 
				if sheet.cell(row=i, column=1).value == student: 
					m = sheet.cell(row=i, column=5).value 
					m = m+1
					sheet.cell(row=i, column=5).value = m+1
					row_num.append(i) 
					no_of_days.append(m) 

	check(no_of_days, row_num, y) 
	resp = int(input('another subject ? 1---->yes 0--->no')) 



























# import openpyxl
# import smtplib
# from email.mime.multipart import MIMEMultipart
# from email.mime.text import MIMEText

# # Constants
# EXCEL_FILE = r'data\\attendance.xlsx'
# EMAIL_SENDER = "your-email@gmail.com"
# EMAIL_PASSWORD = "your-password"
# STAFF_EMAILS = ['bt5161691@gmail.com', 'ambeservices.office@gmail.com']
# SUBJECTS = {1: "CI", 2: "Python", 3: "Data Mining"}

# # Load Excel sheet
# book = openpyxl.load_workbook(EXCEL_FILE)
# sheet = book['Sheet1']

# def save_file():
#     """Save the updated Excel file."""
#     book.save(EXCEL_FILE)
#     print("Attendance file saved.")

# def send_email(to_list, subject, message):
#     """Send an email to the given recipient list."""
#     try:
#         server = smtplib.SMTP('smtp.gmail.com', 587, timeout=120)
#         server.starttls()
#         server.login(EMAIL_SENDER, EMAIL_PASSWORD)

#         msg = MIMEMultipart()
#         msg['Subject'] = subject
#         msg.attach(MIMEText(message, 'plain'))
#         email_content = msg.as_string()

#         for recipient in to_list:
#             server.sendmail(EMAIL_SENDER, recipient, email_content)

#         server.quit()
#         print(f"Mail sent to {len(to_list)} recipients.")

#     except Exception as e:
#         print("Error sending email:", e)

# def process_absentees(subject_id, absentees):
#     """Update attendance and send emails if required."""
#     warning_message = f"Warning! You can take only one more day leave for {SUBJECTS[subject_id]} class."
#     attendance_threshold = 2
#     low_attendance_students = []
#     low_attendance_rolls = []

#     for student_roll in absentees:
#         for row in range(2, sheet.max_row + 1):
#             if sheet.cell(row=row, column=1).value == student_roll:
#                 col = subject_id + 2  # Adjusting column index for attendance
#                 sheet.cell(row=row, column=col).value += 1
#                 save_file()

#                 # Check attendance count
#                 total_leaves = sheet.cell(row=row, column=col).value

#                 if total_leaves == attendance_threshold:
#                     student_email = sheet.cell(row=row, column=2).value
#                     send_email([student_email], "Attendance Warning", warning_message)

#                 elif total_leaves > attendance_threshold:
#                     low_attendance_students.append(sheet.cell(row=row, column=2).value)
#                     low_attendance_rolls.append(str(student_roll))

#     # Notify staff if attendance is too low
#     if low_attendance_students:
#         staff_message = f"The following students have low attendance in {SUBJECTS[subject_id]}: " + ", ".join(low_attendance_rolls)
#         send_email([STAFF_EMAILS[subject_id - 1]], "Lack of Attendance Report", staff_message)
#         send_email(low_attendance_students, "Urgent: Low Attendance", f"You have low attendance in {SUBJECTS[subject_id]}!")

# def main():
#     """Main function to handle attendance updates."""
#     while True:
#         print("1 --> CI\n2 --> Python\n3 --> Data Mining")
#         subject_id = int(input("Enter subject number: "))

#         if subject_id not in SUBJECTS:
#             print("Invalid subject choice.")
#             continue

#         num_absentees = int(input("Number of absentees: "))
#         absentees = list(map(int, input("Enter roll numbers: ").split()))

#         process_absentees(subject_id, absentees)

#         if int(input("Another subject? (1 = Yes, 0 = No): ")) == 0:
#             break

# if __name__ == "__main__":
#     main()
